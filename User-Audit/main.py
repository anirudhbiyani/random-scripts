#!/usr/bin/env python3

from __future__ import print_function
import datetime
import time
import requests
import json
import boto3
import logging.handlers
import os
import re
from openpyxl import Workbook
from slack_sdk import WebClient
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import AuthorizedSession
from oauth2client.service_account import ServiceAccountCredentials

googleusers = {}
hrusers = {}
slackusers = {}
oktausers = {}

def slacknotification():
    color = '#03a9f4'
    if (eventResponse.skipped == True) or (eventResponse.title == "" or eventResponse.author_name == "" or eventResponse.text == ""):        
        print(f"Event: {eventResponse.eventName} SKIPPED.")
        print(f"Event Details:{eventResponse.eventId}")
        
    body = {
            "attachments":[{
                "title": eventResponse.title,
                "author_name": eventResponse.author_name,
                "fallback": eventResponse.author_name,
                "title_link": event_url,
                "text": eventResponse.text,
                "footer": "\nSecurity Alerts | "+str(datetime.now()),
                "color":color,
                "footer_icon": "https://raw.githubusercontent.com/g33kyrash/icons/master/aws_icon_1.png"
        }]}
    if eventResponse.event_type == "GuardDuty Event":
        body["attachments"][0]['footer'] = eventResponse.bottom_text + str(datetime.now()) 
    try:
        headers = {"Content-Type": "application/json"}
        response = requests.post(self.slack_hook_url, headers=headers, json=body)
        print(f"[*]Slack Message sent. Response Code : {response.status_code}")
    except Exception as em:
        print("[-] Error sending slack notification: " + str(em))

def google():
    delegated = ""
    SCOPES = ["https://www.googleapis.com/auth/admin.directory.user.readonly"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name('admin.json', scopes=SCOPES)
    delegated_credentials = credentials.create_delegated(delegated)
    service = build('admin', 'directory_v1', credentials=delegated_credentials)
    
    results = service.users().list(customer='C02kue8sq').execute()
    page = results.get('nextPageToken')
    
    users = results.get('users', [])
    for u in users:
        googleusers[u['primaryEmail'].lower()] = {'name': u['name']['fullName'], 'suspended': u['suspended'], 'archived':  u['archived'], 'lastLogin': u['lastLoginTime']}
        
    while(page):
        results = service.users().list(customer='C02kue8sq', pageToken=page).execute()
        users = results.get('users', [])
        for u in users:
            googleusers[u['primaryEmail'].lower()] = {'name': u['name']['fullName'], 'suspended': u['suspended'], 'archived':  u['archived'], 'lastLogin': u['lastLoginTime']}
        page = results.get('nextPageToken')
    
    return googleusers
    
def hrdata():
    """This function is to read data from a FTE/Contractor/VOP details from a Google Spreadsheet"""

    SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    SPREADSHEET_ID = "10dSfv1ihCjSaNbwbgQbQZ1q7C-1X_A--c3s1EltAQaM"

    creds = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scopes=SCOPES)
    service = build("sheets", "v4", credentials=creds)
    sheet = service.spreadsheets()
    
    # Call the Sheets API
    RANGE_NAME = "FTE!A:F"
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range=RANGE_NAME).execute()
    values = result.get("values", [])
    for i in range(1, len(values)):
        try:
            hrusers[values[i][5].lower().strip()] = {'type': 'FTE', 'status': values[i][0], 'name': values[i][1], 'doj': values[i][2], 'lwd': values[i][3]}
        except IndexError:
            continue
 
    
    RANGE_NAME = "CON + INT!A:F"
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range=RANGE_NAME).execute()
    values = result.get("values", [])
    for i in range(1, len(values)):
        try:
            hrusers[values[i][3].lower().strip()] = {'type': 'CON', 'status': values[i][0], 'name': values[i][2], 'doj': values[i][4], 'lwd': None }
        except IndexError:
            continue

    return hrusers

def okta():
    secret_name = "user-audit/okta-token"
    session = boto3.session.Session()
    client = session.client(service_name='secretsmanager', region_name="ap-south-1")
    get_secret_value_response = client.get_secret_value(SecretId=secret_name)
    token = get_secret_value_response['SecretString']

    headers = {
        "Accept": "application/json",
        'Content-type': 'application/json',
        'Authorization': 'SSWS ' + token
    }

    # The following code blog fetches all the users from Okta and stores them in an array. 
    url = f'https://dreamplug.okta.com/api/v1/users?limit=200'
    response = requests.get(url, headers=headers)
    user = json.loads(response.text)
    for i in range(0,len(user)):
        if user[i]['profile']['secondEmail'] == None:
            oktausers[user[i]['profile']['email'].lower()] = {'firstName': user[i]['profile']['firstName'], 'lastName': user[i]['profile']['lastName'], 'status': user[i]['status'], 'lastLogin': user[i]['lastLogin']}
        else:
            oktausers[user[i]['profile']['secondEmail'].lower()] = {'firstName': user[i]['profile']['firstName'], 'lastName': user[i]['profile']['lastName'], 'status': user[i]['status'], 'lastLogin': user[i]['lastLogin']}

    if "after" in response.headers['link']:
        nextlink = re.search("<(.*?)>", response.headers['link'].split(',')[1])
        nexturl = nextlink.group(0)[:-1][1:]

    while nexturl is not None:
        response = requests.get(nexturl, headers=headers)
        user = json.loads(response.text)
        for i in range(0,len(user)):
            if user[i]['profile']['secondEmail'] == None:
                oktausers[user[i]['profile']['email'].lower()] = {'firstName': user[i]['profile']['firstName'], 'lastName': user[i]['profile']['lastName'], 'status': user[i]['status'], 'lastLogin': user[i]['lastLogin']}
            else:
                oktausers[user[i]['profile']['secondEmail'].lower()] = {'firstName': user[i]['profile']['firstName'], 'lastName': user[i]['profile']['lastName'], 'status': user[i]['status'], 'lastLogin': user[i]['lastLogin']}
        if "next" in response.headers['link']:
            nextlink = re.search(
                "<(.*?)>", response.headers['link'].split(',')[1])
            nexturl = nextlink.group(0)[:-1][1:]
        else:
            nexturl = None

    return oktausers

def slack():
    client = WebClient("<Slack Auth Token>)
    result = client.users_list(team_id='TAWCUK0NT')
    for user in result["members"]:
        try:
            if user["is_bot"] == False:
                if user['deleted'] == 'True':
                    slackusers[user['profile']['email'].lower()] = {'name': user["real_name"], 'status': user['deleted'], 'guest': 'False'}
                elif user['deleted'] == 'False':
                    print(client.users_getPresence(user['profile']['email']))
                    slackusers[user['profile']['email'].lower()] = {'name': user["real_name"], 'status': user['deleted'], 'guest': user['is_restricted']}
        except:
            continue

    while result['response_metadata']['next_cursor']:
        time.sleep(30)
        result = client.users_list(team_id='TAWCUK0NT', cursor = result['response_metadata']['next_cursor'])
        for user in result["members"]:
            try:
                if user["is_bot"] == False:
                    if user['deleted'] == 'True':
                        slackusers[user['profile']['email'].lower()] = {'name': user["real_name"], 'status': user['deleted'], 'guest': 'False'}
                    elif user['deleted'] == 'False':
                        print(client.users_getPresence(user['profile']['email']))
                        slackusers[user['profile']['email'].lower()] = {'name': user["real_name"], 'status': user['deleted'], 'guest': user['is_restricted']}
            except:
                continue

    return slackusers

def main():
    oktausers = okta()
    #googleusers = google()
    slackusers = slack()
    hrusers = hrdata()
    
    wb = Workbook()
    wb.iso_dates = True

    #username | email | deactivation status | LWD | deactivation date | deactivated by

    # Active after Last Working Day
    # Access after LWD in Google Workspace
    glwd = wb.create_sheet('Google - LWD')
    glwd.append(['Name', 'Email Address', 'Status', 'Last Working Day', 'Type'])
    for i in hrusers:
        if hrusers[i]['status'] == 'Inactive' or hrusers[i]['status'] == 'Resigned':
            if hrusers[i]['type'] == "CON" and hrusers[i]['lwd'] == None:
                continue
            elif hrusers[i]['type'] == 'FTE' and hrusers[i]['lwd'] == None:
                print(str(i) + 'Inactive FTE with no defined Last Working Date')
            else:
                try:
                    # There is a bug here - lynal and selva's are still active and this need to be detected here.
                    # Check the variable type
                    # IDK
                    if googleusers[i]['suspended'] == False:
                        glwd.append([googleusers[i]['name'], i, 'Active' , hrusers[i]['lwd'], hrusers[i]['type']])
                except KeyError:
                    print('Could not find them in Google Workspace - ' + str(i) + ' - ' + hrusers[i]['lwd'])    

    # Access after LWD in Okta
    olwd = wb.create_sheet('Okta - LWD')
    olwd.append(['Name', 'Email Address', 'Status', 'Last Working Day', 'Type'])
    for i in hrusers:
        if hrusers[i]['status'] == 'Inactive' or hrusers[i]['status'] == 'Resigned':
            if hrusers[i]['type'] == 'CON' and hrusers[i]['lwd'] == None:
                continue
            elif hrusers[i]['type'] == 'FTE' and hrusers[i]['lwd'] == None:
                print(str(i) + 'Inactive FTE with no defined Last Working Date')
            else:
                try:
                    if oktausers[i]['status'] == 'ACTIVE':
                        olwd.append([oktausers[i]['firstName'] + ' ' + oktausers[i]['lastName'], i, oktausers[i]['status'], hrusers[i]['lwd'], hrusers[i]['type']])
                except KeyError:
                    print('Could not find them in Okta - ' + str(i) + ' - ' + hrusers[i]['lwd'])    
    
    # This is not working and needs to be fixed
    # Access after LWD in Slack
    slwd = wb.create_sheet('Slack - LWD')
    slwd.append(['Name', 'Email Address', 'Status', 'Last Working Day', 'Type'])
    print(slackusers)
    for i in hrusers:
        if hrusers[i]['status'] == 'Inactive' or hrusers[i]['status'] == 'Resigned':
            if hrusers[i]['type'] == 'CON' and hrusers[i]['lwd'] == None:
                continue
            elif hrusers[i]['type'] == 'FTE' and hrusers[i]['lwd'] == None:
                print(str(i) + 'Inactive FTE with no defined Last Working Date')
            else:
                try:
                    print(slackusers[i])
                    print(type(slackusers[i]))
                    if slackusers[i]['status'] == True:  
                        print(slackusers[i]['name'], i, slackusers[i]['status'], hrusers[i]['lwd'], hrusers[i]['type'])
                        slwd.append(slackusers[i]['name'], i, slackusers[i]['status'], hrusers[i]['lwd'], hrusers[i]['type'])
                except KeyError:
                    print('Could not find them in Slack - ' + str(i) + ' - ' + hrusers[i]['lwd']) 
    
    # All the users who have not logged in last 90 days
    # Google Workspace
    google30 = wb.create_sheet('Google - Inactive Logins')
    google30.append(['Name', 'Email Address', 'Status', 'LastLogin'])
    delta = datetime.datetime.now() - datetime.timedelta(days=30)
    for i in googleusers:
        if(googleusers[i]['lastLogin'] == None):
            # print(str(i) + ' has never logged into Google')
            google30.append([googleusers[i]['name'], i, googleusers[i]['suspended'] , 'No Login Found'])
        elif (delta > datetime.datetime.strptime(googleusers[i]['lastLogin'].split('T')[0], '%Y-%m-%d')):
            # print(str(i) + ' has not logged into Google for more than 90 days. Last Login - ' + str(datetime.datetime.strptime(googleusers[i]['lastLogin'].split('T')[0], '%Y-%m-%d')))
            google30.append([googleusers[i]['name'], i, googleusers[i]['suspended'] , datetime.datetime.strptime(googleusers[i]['lastLogin'].split('T')[0], '%Y-%m-%d')])
    
    # Okta
    delta = datetime.datetime.now() - datetime.timedelta(days=30)
    okta30 = wb.create_sheet('Okta - Inactive Logins')
    okta30.append(['Name', 'Email Address', 'Status', 'LastLogin'])
    for i in oktausers:
        if(oktausers[i]['lastLogin'] == None):
            print(oktausers[i])
            okta30.append([ oktausers[i]['firstName']  + ' ' +  oktausers[i]['lastName'], i, oktausers[i]['status'] , 'No Login Found'])
        elif (delta > datetime.datetime.strptime(oktausers[i]['lastLogin'].split('T')[0], '%Y-%m-%d')):
            # print(str(i) + ' has not logged into Okta for more than 90 days. Last Login - ' + str(datetime.datetime.strptime(oktausers[i]['lastLogin'].split('T')[0], '%Y-%m-%d')))
            okta30.append([ oktausers[i]['firstName'] + ' ' + oktausers[i]['lastName'] , i, oktausers[i]['status'] , datetime.datetime.strptime(oktausers[i]['lastLogin'].split('T')[0], '%Y-%m-%d')])
    
    # List of users who are present in IT but not in HR system.
    # Google Workspace
    gextra = wb.create_sheet('Google - Extra Users')
    gextra.append(["Name", "Email", "Status", "Last Login"])
    extragoogle = set(googleusers) - set(hrusers)
    for i in extragoogle:
        try:
            if googleusers[i]['suspended'] == False:  
                gextra.append([googleusers[i]['name'], i, 'Active', googleusers[i]['lastLogin']])
        except KeyError:
            continue

    # Okta
    oextra = wb.create_sheet('Okta - Extra Users')
    oextra.append(["Name", "Email", "Status", "Last Login"])
    extraokta = set(oktausers) - set(hrusers)
    for i in extraokta:
        try:
            if oktausers[i]['status'] == 'ACTIVE':  
                if(oktausers[i]['lastLogin'] == None):
                    oextra.append([ oktausers[i]['firstName'] + ' ' + oktausers[i]['lastName'] , i, 'Active', 'No Logins'])
                else:
                    oextra.append([ oktausers[i]['firstName'] + ' ' + oktausers[i]['lastName'] , i, 'Active', datetime.datetime.strptime(oktausers[i]['lastLogin'].split('T')[0], '%Y-%m-%d')])
        except KeyError:
            continue
    
    #Slack
    sextra = wb.create_sheet('Slack - Extra Users')
    sextra.append(["Name", "Email", "Status", "Guest"])
    extraslack = set(slackusers) - set(hrusers)
    for i in extraslack:
        try:
            if slackusers[i]['status'] == False:
                sextra.append([slackuser[i]['name'], i, 'Active', slackuser[i]['guest']])
        except KeyError:
            continue
    
    wb.save(filename = 'user-audit-' + str(datetime.date.today())+'.xlsx')
    
if __name__ == "__main__":
    main()
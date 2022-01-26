#!/usr/bin/env python

import openpyxl, getpass, sys, os
from ldap3 import Server, Connection, ALL_ATTRIBUTES, ALL, SUBTREE, AUTO_BIND_NO_TLS

"""
Before running this script please make sure you have openpyxl and ldap3 libraries installed on your box.
"""

__author__ = "Aniruddha Biyani"
__version__ = "1.0.0"
__maintainer__ = "Aniruddha Biyani"
__email__ = "aniruddha.biyani@lithium.com"
__status__ = "InfoSec_Utilities"

notPresent =[]
activeEmail=[]
disabledEmail=[]
otherEmail=[]
dlEmail=[]
conn = 0

def ldapSearch( emailAddress):
    global notPresent
    global activeEmail
    global disabledEmail
    global otherEmails
    global dlEmail
    global conn

    base_dn="DC=lithium, DC=local"
    user = emailAddress.split("@")[0]
    conn.search(search_base=base_dn, search_filter='(sAMAccountName='+user+')', search_scope=SUBTREE, attributes='userAccountControl')
    result = conn.entries
    if( len(result) == 0 ):
        notPresent.append(emailAddress)
    else:
        if 'OU=Distro Groups' in str(result[0]):
            dlEmail.append(emailAddress)
        else:
            print result[0]
            responseCode = str(result[0]).split("userAccountControl: ")[1].strip()
            if(responseCode == '512'):
                activeEmail.append(emailAddress)
            elif(responseCode == '514'):
                disabledEmail.append(emailAddress)
            else:
                otherEmail.append(emailAddress)

def ldapConnect():
    global conn
    user = raw_input("Enter your username that you want to use to connect: ")
    pswd = getpass.getpass('Password:')
    usr="lithium.local\\"+user

    base_dn="OU=Users, OU=Lithium, DC=lithium, DC=local"

    srv = Server('idcdc01.lithium.local', use_ssl=True, get_info=ALL)
    conn = Connection(srv, user=usr, password=pswd, authentication="NTLM")
    conn.bind()

def excel(file):
    wb = openpyxl.load_workbook(file)
    pwnedemail = wb['Breached email accounts']
    pastedemail = wb['Pasted email accounts']
    emailList = []
    for row in pwnedemail.iter_rows('A{}:A{}'.format(pwnedemail.min_row,pwnedemail.max_row)):
        for cell in row:
            emailList.append(cell.value)

    for row in pastedemail.iter_rows('A{}:A{}'.format(pastedemail.min_row,pastedemail.max_row)):
        for cell in row:
            emailList.append(cell.value)

    return emailList

def main():
    global notPresent
    global activeEmail
    global disabledEmail
    global otherEmail
    global conn

    email = []

    if os.path.isfile(sys.argv[1]):
        if(len(sys.argv) != 2):
            print("Invalid Number of Arguments.")
            exit()
        else:
            email = excel(sys.argv[1])
            ldapConnect()
    else:
        print "File does not exists."

    for i in email:
        if(i=="Email"):
            continue
        else:
            ldapSearch(i)

    print("-------------------------------------------------------------------------------------------------------------------------------------")
    print "\n".join(notPresent)
    print("-------------------------------------------------------------------------------------------------------------------------------------")
    print "\n".join(activeEmail)
    print("-------------------------------------------------------------------------------------------------------------------------------------")
    print "\n".join(disabledEmail)
    print("-------------------------------------------------------------------------------------------------------------------------------------")
    print "\n".join(otherEmail)
    print("-------------------------------------------------------------------------------------------------------------------------------------")
    print "\n".join(dlEmail)
    print("-------------------------------------------------------------------------------------------------------------------------------------")

if __name__ == '__main__':
    main()

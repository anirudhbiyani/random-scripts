#!/usr/bin/env python

import openpyxl, getpass, sys, os, hashlib, datetime
from ldap3 import Server, Connection, ALL_ATTRIBUTES, ALL, SUBTREE, AUTO_BIND_NO_TLS

"""
Before running this script please make sure you have openpyxl and ldap3 libraries installed on your box.
"""

__author__ = "Aniruddha Biyani"
__version__ = "1.0.0"
__maintainer__ = "Aniruddha Biyani"
__email__ = "contact@anirudbiyani.com"
__status__ = "Active"

def main():
    base_dn="DC=test, DC=local"
    domain="test.local"
    domainController=""
    notFound = []
    inSLA = []
    outSLA = []

# Making sure that the fietype is correct and the number of arguments is proper.
    if(len(sys.argv) != 2):
        print("Invalid Number of Arguments.")
        print("Syntax - {} Path to the file".format(sys.argv[0]))
        exit(1)

    if os.path.isfile(sys.argv[1]):
        if (sys.argv[1].endswith('.xlsx') or sys.argv[1].endswith('.xls')):
            pass
    else:
        print "File does not exists or is not the right filetype."
        exit(1)

    user = raw_input("Enter your username that you want to use to connect: ")
    pswd = getpass.getpass('Password:')
    usr= domain+"\\"+user
    srv = Server(domainController', use_ssl=True, get_info=ALL)
    conn = Connection(srv, user=usr, password=pswd, authentication="NTLM")
    conn.bind()

# Loading the excel file into the program
    wb = openpyxl.load_workbook(sys.argv[1], read_only=True, data_only=True)
    worksheet = wb['Sheet1']
    count = worksheet.max_row + 1

    print '-' * 30

    for i in range(2, count):
        firstName = worksheet.cell(i,2).value
        lastName = worksheet.cell(i,3).value
        termDate = worksheet.cell(i,6).value
        positionID = worksheet.cell(i,1).value
        title = worksheet.cell(i,8).value
        deptCode =  worksheet.cell(i,9).value
        deptName = worksheet.cell(i,10).value
        location =  worksheet.cell(i,12).value

        termDate = termDate.date()

        employee = (firstName + '.' + lastName).lower()
        semployee = (firstName[0] + lastName).lower()
        conn.search(search_base=base_dn, search_filter='(sAMAccountName='+employee+')', search_scope=SUBTREE, attributes=['accountExpires', 'userAccountControl', 'pwdLastSet', 'lastLogonTimestamp'])
        aresult = conn.entries

        conn.search(search_base=base_dn, search_filter='(sAMAccountName='+semployee+')', search_scope=SUBTREE, attributes=['accountExpires', 'userAccountControl', 'pwdLastSet', 'lastLogonTimestamp'])
        sresult = conn.entries

        if (len(aresult) == 0) and (len(sresult) == 0):
            notFound.append(firstName + ' ' + lastName)
        else:
            if not aresult:
                result = sresult
            elif not sresult:
                result = aresult
            else:
                print'Somethingn is wrong with the code'

            pswdChange = str(result[0]).split("pwdLastSet: ")[1].split('\n')[0].strip()
            pswdChange = datetime.datetime.strptime(pswdChange, "%Y-%m-%d %H:%M:%S.%f+00:00").date()

            if "lastLogonTimestamp" not in result:
                loginTime = pswdChange
            else:
                loginTime = str(result[0]).split("lastLogonTimestamp: ")[1].split('\n')[0].strip()
                loginTime = datetime.datetime.strptime(loginTime, "%Y-%m-%d %H:%M:%S.%f+00:00").date()

            if (loginTime < pswdChange) or (loginTime == pswdChange):
                a = 1
            else:
                a = 0

            if ((termDate + datetime.timedelta(days=1)) == pswdChange) or (termDate == pswdChange) or ((termDate + datetime.timedelta(days=2)) == pswdChange) or (termDate > pswdChange):
                b = 1
            else:
                b = 0

            if (termDate - pswdChange > datetime.timedelta(days=7)):
                b = 0
            else:
                b = 1

            if ((str(result[0]).split("userAccountControl: ")[1].strip() == '514') and (a == 1) and (b == 1)):
                inSLA.append(firstName + ' ' + lastName)
            else:
                outSLA.append(firstName + ' ' + lastName)

    conn.unbind()

    print 'The following users were disabled within SLA -'
    print("\n".join(inSLA))
    print '-' * 30
    print 'The following users were not disabled within SLA -'
    print("\n".join(outSLA))
    print '-' * 30
    print 'The following users were not found in the Active Directory -'
    print("\n".join(notFound))

if __name__ == '__main__':
    main()

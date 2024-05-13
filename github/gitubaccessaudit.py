#!/usr/bin/env python

from github import Github
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook

__author__ = "Aniruddha Biyani"
__version__ = "1.0.0"
__maintainer__ = "Aniruddha Biyani"
__email__ = "aniruddha.biyani@lithium.com"
__status__ = "InfoSec_Utilities"

def main():
    g = Github('<token>')

    wb = Workbook()
    rl = g.get_rate_limit()
    if(rl.core.remaining == 1):
        sleep(3700)

    for r in repo0:
        ro = g.get_repo(str(r))
        print r
        i = 2
        ws = wb.create_sheet(title=ro.name)
        collaborators = ro.get_collaborators()
        for collaborator in collaborators:
            print collaborator
            name = collaborator.name
            login = collaborator.login
            perm = ro.get_collaborator_permission(login)
            ws['A1'] = 'Name'
            ws['B1'] = 'GitHub User ID'
            ws['C1'] = 'Permission'
            ws['A' + str(i)] = name
            ws['B' + str(i)] = login
            ws['C' + str(i)] = perm
            i = i + 1
            rl = g.get_rate_limit()
            if(rl.core.remaining == 1)
                sleep(3700)

    wb.remove(xlsx['Sheet'])
    wb.remove(xlsx['Sheet1'])
    wb.save(filename = 'access_audit_github.xlsx')

if __name__ == '__main__':
    main()

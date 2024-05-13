#!/usr/bin/env python

from github import Github
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook

__author__ = "Aniruddha Biyani"
__version__ = "1.0.0"
__maintainer__ = "Aniruddha Biyani"
__email__ = "contact@anirudbiyani.com"
__status__ = "Active"

def main():
    token = '' #Your Github Token
    g = Github(token)
    wb = Workbook()
    repo = [] # List of repos for which you want to get permissions for.

    rl = g.get_rate_limit()
    if(rl.core.remaining == 1):
        sleep(3700)

    for r in repo:
        ro = g.get_repo(str(r))
        print r
        i = 2
        ws = wb.create_sheet(title=ro.name)
        collaborators = ro.get_collaborators()
        for collaborator in collaborators:
            print collaborator
            name = collaborator.name
            login = collaborator.login
            perm = ro.get_collaborator_permission(login)
            ws['A1'] = 'Name'
            ws['B1'] = 'GitHub User ID'
            ws['C1'] = 'Permission'
            ws['A' + str(i)] = name
            ws['B' + str(i)] = login
            ws['C' + str(i)] = perm
            i = i + 1
            rl = g.get_rate_limit()
            if(rl.core.remaining == 1)
                sleep(3700)

    xlsx.remove(xlsx['Sheet'])
    xlsx.remove(xlsx['Sheet1'])
    xlsx.save(filename = 'access_audit_github_repo.xlsx')

if __name__ == '__main__':
    main()
